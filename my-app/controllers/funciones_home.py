
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


def insertarsegro(valor1):
    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO seguro_salud (Nombre) VALUES (%s)"
                valores = (valor1,)
                cursor.execute(sql, valores)
                resultado_insert = cursor.lastrowid
                conexion_MySQLdb.commit()


                print('last id :', resultado_insert)
        return resultado_insert

    except Exception as e:
        print('last errorrr xx :', cursor.lastrowid)
        return f'Se produjo un error en xxxxxxxxxxxxxxxxxxxxxxxxxxx: {str(e)}'

def procesar_form_empleado(dataForm):
    # Formateando Salario
    #salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['Sueldo'])  

    Sueldo = int(dataForm['Sueldo'])
    Nombre_Completo = dataForm['Nombre_Completo']
    RUT = dataForm['RUT']
    Sexo = dataForm['Sexo']
    Direccion = dataForm['Direccion']
    Telefono = dataForm['Telefono']
    #Fecha_Ingreso = dataForm['Fecha_Ingreso']
    Cargo = dataForm['Cargo']
    Id_Seguro_Salud = int(dataForm['Id_Seguro_Salud'])
    Id_Region = int( dataForm['Id_Region'])
    Id_comuna = 2
    Rol = int(dataForm['Id_Rol'])
    email_user = dataForm['email_user']
    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO trabajador (Nombre_Completo, RUT, Sexo, Direccion, Telefono, Cargo, Id_Seguro_Salud, Id_Region, Id_comuna, Rol, email_user, Sueldo) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                # Creando una tupla con los valores del INSERT
                valores = (
                 Nombre_Completo
                ,RUT
                ,Sexo
                ,Direccion
                ,Telefono
                # ,Fecha_Ingreso
                ,Cargo
                , Id_Seguro_Salud
                , Id_Region
                , Id_comuna
                ,Rol
                ,email_user
                ,Sueldo
                )
                cursor.execute(sql, valores)
                resultado_insert = cursor.lastrowid
                #resultado_insert = cursor.rowcount
                
                sql2 = "INSERT INTO cargas_familiares (Id_Trabajador, Nombre_Carga, Parentesco, Sexo, RUT_Carga) VALUES (%s, %s, %s, %s, %s)"
                valores2 = (
                resultado_insert
                ,dataForm['Nombre_Carga']
                ,dataForm['Parentesco']
                ,dataForm['Sexo']
                ,dataForm['RUT_Carga']
                )
                cursor.execute(sql2, valores2)               
                resultado_insert2 = cursor.lastrowid

                sql3 = "INSERT INTO contacto_emergencia (Id_Trabajador, Nombre_Contacto, Relacion, Telefono_Contacto) VALUES (%s, %s, %s, %s)"
                valores3 = (
                resultado_insert
                ,dataForm['Nombre_Contacto']
                ,dataForm['Relacion']
                ,dataForm['Telefono_Contacto']
                )
                cursor.execute(sql3, valores3)               
                filaretornada = cursor.rowcount
                
                conexion_MySQLdb.commit()
        return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'

def sql_inserta_cargasfamiliares(dataForm, idTrabajador):
    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO cargas_familiares (Id_Trabajador, Nombre_Carga, Parentesco, Sexo, RUT_Carga) VALUES (%s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (
                idTrabajador
                ,dataForm['Nombre_Carga']
                ,dataForm['Parentesco']
                ,dataForm['Sexo']
                ,dataForm['RUT_Carga']
                )
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                resultado_insert = cursor.lastrowid
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_cargas familiares: {str(e)}'

# Lista de Empleados que se muestra en la tabla con botones para editar, actualizar y eliminar
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.Id_Trab,
                        e.RUT,
                        e.Nombre_Completo,
                        CASE
                            WHEN e.Sexo = '1' THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS Sexo,
                        e.Cargo,
                        e.email_user,
                        e.Rol
                    FROM trabajador AS e
                    ORDER BY e.Id_Trab DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.Id_Trab,
                        e.Nombre_Completo, 
                        e.RUT,
                        CASE
                            WHEN e.Sexo = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS Sexo,
                        e.Direccion, 
                        e.Telefono,
                        e.Cargo,
                        e.Sueldo
                                 
                    FROM trabajador AS e
                    WHERE Id_Trab =%s
                    ORDER BY e.Id_Trab DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.RUT,
                        e.Nombre_Completo, 
                        CASE
                            WHEN e.Sexo = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS Sexo,
                        e.Telefono,
                        e.email_user,
                        e.Cargo,
                        e.Direccion,
                        e.Sueldo
                        
                    FROM trabajador AS e
                    ORDER BY e.Id_Trab DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()

    print('dataEmpleados :::: ', dataEmpleados)

    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("RUT", "Nombre competo", "Sexo",
                     "Telefono", "Correo electrónico", "Cargo", "Direccion", "Sueldo")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        RUT = registro['RUT']
        Nombre_Completo = registro['Nombre_Completo']
        Sexo = registro['Sexo']
        Telefono = registro['Telefono']
        email_user = registro['email_user']
        Cargo = registro['Cargo']
        Direccion = registro['Direccion']
        Sueldo = registro['Sueldo']

        # Agregar los valores a la hoja
        hoja.append((RUT, Nombre_Completo, Sexo, Telefono, email_user, Cargo,
                     Direccion, Sueldo))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        select t.Id_Trab, t.Nombre_Completo, t.RUT,t.Sexo, t.Direccion, t.Telefono, t.Cargo, t.rol
                        ,t.Sueldo, t.email_user, cf.Id_Carga, cf.Nombre_Carga, cf.Parentesco, cf.Sexo as SexoCarga, cf.RUT_Carga
                        ,ce.Id_Contacto, ce.Nombre_Contacto, ce.Relacion, ce.Telefono_Contacto 
                        from trabajador t 
                        inner join cargas_familiares cf on t.Id_Trab = cf.Id_Trabajador
                        inner join contacto_emergencia ce on ce.Id_Trabajador = t.Id_Trab
                        WHERE t.Id_Trab =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []

# Trae los datos del empleado logeado para actualizar sus datos desde su perfil menu
def info_datos_personales(email):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        select t.Id_Trab, t.Nombre_Completo, t.RUT,t.Sexo, t.Direccion, t.Telefono, t.Cargo, t.rol
                        ,t.Sueldo, t.email_user, cf.Id_Carga, cf.Nombre_Carga, cf.Parentesco, cf.Sexo as SexoCarga, cf.RUT_Carga
                        ,ce.Id_Contacto, ce.Nombre_Contacto, ce.Relacion, ce.Telefono_Contacto 
                        from trabajador t 
                        inner join cargas_familiares cf on t.Id_Trab = cf.Id_Trabajador
                        inner join contacto_emergencia ce on ce.Id_Trabajador = t.Id_Trab
                        WHERE t.email_user =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (email,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []

# SOLO PUEDE ACTUALIZAR EL EMPLEADO DE RRHH - ROL 2              
def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # salario_sin_puntos = re.sub(
                #     '[^0-9]+', '', data.form['salario_empleado'])
                # salario_empleado = int(salario_sin_puntos)
                
                Id_Contacto = data.form['Id_Contacto']
                Nombre_Contacto = data.form['Nombre_Contacto']
                Relacion = data.form['Relacion']
                Telefono_Contacto = data.form['Telefono_Contacto']  
                Id_Carga = data.form['Id_Carga']   
                Nombre_Carga = data.form['Nombre_Carga'] 
                RUT_Carga = data.form['RUT_Carga'] 
                Parentesco = data.form['Parentesco']   
                SexoCarga = data.form['SexoCarga']       
                querySQL = """
                    UPDATE contacto_emergencia
                    SET 
                        Nombre_Contacto = %s,
                        Relacion = %s,
                        Telefono_Contacto = %s
                    WHERE Id_Contacto = %s
                """
                values = (Nombre_Contacto, Relacion, Telefono_Contacto, Id_Contacto)
                cursor.execute(querySQL, values)

                querySQL2 = """
                    UPDATE cargas_familiares
                    SET 
                        Nombre_Carga = %s,
                        Parentesco = %s,
                        Sexo = %s,
                        RUT_Carga = %s
                    WHERE Id_Carga = %s
                """
                values2 = (Nombre_Carga, Parentesco, SexoCarga, RUT_Carga, Id_Carga)
                cursor.execute(querySQL2, values2)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None

# EL TRABAJADOR ACTUALIZA ALGUNOS DATOS - ROL 3
def procesar_actualizacion_formTrabajador(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                Id_Contacto = data.form['Id_Contacto']
                Nombre_Contacto = data.form['Nombre_Contacto']
                Relacion = data.form['Relacion']
                Telefono_Contacto = data.form['Telefono_Contacto']  
                Id_Carga = data.form['Id_Carga']   
                Nombre_Carga = data.form['Nombre_Carga'] 
                RUT_Carga = data.form['RUT_Carga'] 
                Parentesco = data.form['Parentesco']   
                SexoCarga = data.form['SexoCarga']       
                querySQL = """
                    UPDATE contacto_emergencia
                    SET 
                        Nombre_Contacto = %s,
                        Relacion = %s,
                        Telefono_Contacto = %s
                    WHERE Id_Contacto = %s
                """
                values = (Nombre_Contacto, Relacion, Telefono_Contacto, Id_Contacto)
                cursor.execute(querySQL, values)

                querySQL2 = """
                    UPDATE cargas_familiares
                    SET 
                        Nombre_Carga = %s,
                        Parentesco = %s,
                        Sexo = %s,
                        RUT_Carga = %s
                    WHERE Id_Carga = %s
                """
                values2 = (Nombre_Carga, Parentesco, SexoCarga, RUT_Carga, Id_Carga)
                cursor.execute(querySQL2, values2)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado SOLO PUEDE ELIMINAR EL ROL 2 (empleado rrhh)
def eliminarEmpleado(id_empleado, email_user):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                querySQL = "DELETE FROM trabajador WHERE Id_Trab=%s"
                cursor.execute(querySQL, (id_empleado,))

                querySQL2 = "DELETE FROM users WHERE email_user=%s"
                cursor.execute(querySQL2, (email_user,))

                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount               
        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []


# get seguros de salud
def sql_lista_segurosSaludBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.Id_Seguro_Salud,
                        e.Nombre
                    FROM seguro_salud AS e
                    """)
                cursor.execute(querySQL,)
                sueldosBD = cursor.fetchall()
        return sueldosBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_segurosSaludBD: {e}")
        return None

# get regiones
def sql_lista_regionesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.Id_Region,
                        e.Nombre
                    FROM region AS e
                    """)
                cursor.execute(querySQL,)
                sueldosBD = cursor.fetchall()
        return sueldosBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_regionesBD: {e}")
        return None

# get roles
def sql_lista_rolesBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.Id_Rol,
                        e.Rol
                    FROM rol AS e
                    """)
                cursor.execute(querySQL,)
                sueldosBD = cursor.fetchall()
        return sueldosBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_rolesBD: {e}")
        return None